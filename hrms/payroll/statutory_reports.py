# Copyright (c) 2026, Frappe Technologies Pvt. Ltd. and contributors
# For license information, please see license.txt

"""Statutory (government) payroll returns generated from an approved Payroll Entry.

Finance used to fill these workbooks by hand every month. Each report is the
authority's own template with the payroll's employee rows written into it.
"""

import os
import re
from io import BytesIO

import openpyxl

import frappe
from frappe import _
from frappe.utils import flt, getdate

TEMPLATE_DIR = os.path.join(os.path.dirname(__file__), "statutory_templates")

REPORTS = ("PAYE", "SDL", "NSSF", "NHIF", "WCF")

# component name -> value we pull off each salary slip
BASIC_COMPONENT = "Basic"
NSSF_COMPONENT = "NSSF"


def digits(value: str) -> str:
	"""Authority portals reject TINs/IDs with separators."""
	return re.sub(r"\D", "", str(value or ""))


def get_rows(payroll_entry: str) -> list[dict]:
	"""Employee + salary slip figures for every slip in the payroll entry."""
	ss = frappe.qb.DocType("Salary Slip")
	emp = frappe.qb.DocType("Employee")

	rows = (
		frappe.qb.from_(ss)
		.join(emp)
		.on(ss.employee == emp.name)
		.select(
			ss.name.as_("salary_slip"),
			emp.name.as_("employee"),
			emp.employee_name,
			emp.first_name,
			emp.middle_name,
			emp.last_name,
			emp.gender,
			emp.date_of_birth,
			emp.date_of_joining,
			emp.marital_status,
			emp.company_email,
			emp.personal_email,
			emp.cell_number,
			emp.designation,
			emp.employment_type,
			emp.tin_number,
			emp.nssf_number,
			emp.nhif_number,
			emp.wcf_number,
			emp.nida_number,
		)
		.where((ss.payroll_entry == payroll_entry) & (ss.docstatus != 2))
		.orderby(emp.employee_name)
	).run(as_dict=True)

	if not rows:
		frappe.throw(_("No salary slips found for {0}").format(payroll_entry))

	amounts = frappe.get_all(
		"Salary Detail",
		filters={
			"parent": ("in", [r.salary_slip for r in rows]),
			"salary_component": ("in", [BASIC_COMPONENT, NSSF_COMPONENT]),
		},
		fields=["parent", "salary_component", "amount"],
	)
	by_slip = {}
	for a in amounts:
		by_slip.setdefault(a.parent, {})[a.salary_component] = flt(a.amount)

	for r in rows:
		slip = by_slip.get(r.salary_slip, {})
		r.basic = slip.get(BASIC_COMPONENT, 0)
		r.nssf = slip.get(NSSF_COMPONENT, 0)

	return rows


def load_template(filename: str):
	return openpyxl.load_workbook(os.path.join(TEMPLATE_DIR, filename))


def write(ws, rows: list[list], start_row: int = 2, clear_until: int = 3):
	"""Drop the template's sample rows, then write ours from start_row."""
	for row in ws.iter_rows(min_row=start_row, max_row=max(clear_until, ws.max_row)):
		for cell in row:
			cell.value = None

	for i, values in enumerate(rows):
		for j, value in enumerate(values):
			ws.cell(row=start_row + i, column=j + 1, value=value)


def as_download(wb, filename: str):
	out = BytesIO()
	wb.save(out)
	frappe.response["filename"] = filename
	frappe.response["filecontent"] = out.getvalue()
	frappe.response["type"] = "binary"


def build_paye(rows, doc):
	wb = load_template("paye.xlsx")
	wb["INSTRUCTIONS"]["B4"] = frappe.db.get_value("Company", doc.company, "tax_id") or ""
	write(
		wb["PAYE_EMPLOYEE"],
		[
			[
				i + 1,
				digits(r.tin_number),
				(r.employee_name or "").upper(),
				digits(r.nida_number),
				"Primary",
				"Resident",
				r.nssf_number or "",
				r.basic,
				0,  # allowance: excluded until finance confirms the components
				0,  # benefits in kind
				r.nssf,
				"Tanzania Mainland",
			]
			for i, r in enumerate(rows)
		],
	)
	return wb


def build_sdl(rows, doc):
	wb = load_template("sdl.xlsx")
	end = getdate(doc.end_date)
	instructions = wb["INSTRUCTIONS"]
	instructions["B3"] = digits(frappe.db.get_value("Company", doc.company, "tax_id"))
	instructions["B4"] = str(end.year)
	instructions["B5"] = end.strftime("%B")
	write(
		wb["Tanzania Mainland"],
		[
			[
				i + 1,
				digits(r.tin_number),
				(r.employee_name or "").upper(),
				digits(r.nida_number),
				"Permanent",
				"Resident",
				r.basic,
				0,  # allowance
				0,  # exempted amount
				"Tanzania Mainland",
			]
			for i, r in enumerate(rows)
		],
	)
	return wb


def build_nssf(rows, doc):
	wb = load_template("nssf.xlsx")
	write(
		wb["Template"],
		[
			[
				i + 1,
				r.nssf_number or "",
				r.first_name or "",
				r.middle_name or "",
				r.last_name or "",
				r.basic,
			]
			for i, r in enumerate(rows)
		],
		clear_until=2,
	)
	return wb


def build_wcf(rows, doc):
	wb = load_template("wcf.xlsx")
	write(
		wb["mySheet"],
		[
			[
				r.wcf_number or "",
				(r.first_name or "").upper(),
				(r.middle_name or "").upper(),
				(r.last_name or "").upper(),
				(r.gender or "").upper(),
				r.date_of_birth,
				r.basic,
				r.basic,
				r.designation or "",
				r.employment_type or "",
				digits(r.nida_number),
				r.cell_number or "",
			]
			for r in rows
		],
	)
	return wb


NHIF_COLUMNS = [
	"First Name",
	"Middle Name",
	"Last Name",
	"Date Of Birth",
	"Gender",
	"Marital Status",
	"Email Address",
	"Mobile Number",
	"National ID",
	"WCF Number",
	"Recruitment Date",
	"Basic Salary",
]


def build_nhif(rows, doc):
	# the authority's own file is legacy .xls, which we cannot write; same columns, xlsx
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "Employees List"
	ws.append(NHIF_COLUMNS)
	for r in rows:
		ws.append(
			[
				(r.first_name or "").upper(),
				(r.middle_name or "").upper(),
				(r.last_name or "").upper(),
				r.date_of_birth,
				r.gender or "",
				r.marital_status or "",
				r.company_email or r.personal_email or "",
				r.cell_number or "",
				digits(r.nida_number),
				r.wcf_number or "",
				r.date_of_joining,
				r.basic,
			]
		)
	return wb


BUILDERS = {
	"PAYE": build_paye,
	"SDL": build_sdl,
	"NSSF": build_nssf,
	"NHIF": build_nhif,
	"WCF": build_wcf,
}


@frappe.whitelist()
def download(payroll_entry: str, report: str):
	if report not in BUILDERS:
		frappe.throw(_("Unknown statutory report {0}").format(report))

	doc = frappe.get_doc("Payroll Entry", payroll_entry)
	doc.check_permission("read")

	wb = BUILDERS[report](get_rows(payroll_entry), doc)
	as_download(wb, f"{report}-{payroll_entry}.xlsx")
